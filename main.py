import openpyxl
from eth_account import Account
from bip_utils import Bip39MnemonicGenerator, Bip39SeedGenerator, Bip44, Bip44Coins, Bip44Changes


def generate_wallets(num_wallets):
    wallets = []

    for wallet in range(num_wallets):
        mnemonic_generator = Bip39MnemonicGenerator()
        mnemonic = mnemonic_generator.FromWordsNumber(words_num=12)
        seed = Bip39SeedGenerator(mnemonic).Generate()
        bip_obj = Bip44.FromSeed(seed, Bip44Coins.ETHEREUM)

        private_key = bip_obj.Purpose().Coin().Account(0).Change(Bip44Changes.CHAIN_EXT).AddressIndex(0).PrivateKey().Raw().ToBytes()
        account = Account.from_key(private_key)

        wallet = {
            'address': account.address,
            'private_key': private_key.hex(),
            'mnemonic': mnemonic,
        }

        wallets.append(wallet)

    return wallets


def save_wallets_to_xlsx(wallets, file_name):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Mnemonic"
    ws['B1'] = "Address"
    ws['C1'] = "Private Key"

    for row, wallet in enumerate(wallets, start=2):
        ws.cell(row=row, column=1, value=str(wallet['mnemonic']))
        ws.cell(row=row, column=2, value=wallet['address'])
        ws.cell(row=row, column=3, value=wallet['private_key'])

    # Save the workbook
    wb.save(file_name)


def save_wallets_to_files(wallets):
    with open('seeds.txt', 'w') as f:
        for wallet in wallets:
            f.write(f"{wallet['mnemonic']}\n")

    with open('addresses.txt', 'w') as f:
        for wallet in wallets:
            f.write(f"{wallet['address']}\n")

    with open('private_key.txt', 'w') as f:
        for wallet in wallets:
            f.write(f"{wallet['private_key']}\n")


if __name__ == '__main__':
    num_wallets = int(input("Введите количество кошельков для генерации: "))
    wallets = generate_wallets(num_wallets)
    save_wallets_to_xlsx(wallets, 'wallets.xlsx')
    save_wallets_to_files(wallets)
    print(f"{num_wallets} кошельков сгенерировано и сохранено в файлах")
